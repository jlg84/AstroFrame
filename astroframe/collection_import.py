from __future__ import annotations

import math
import re
from pathlib import Path
from typing import Any

from .knowledge import (
    CollectionEntry,
    CollectionRecord,
    KnowledgeStore,
    TargetRecord,
    slugify,
    target_id_for,
)


def _text(value: Any) -> str | None:
    if value is None:
        return None
    value = str(value).strip()
    return value or None


def _float(value: Any) -> float | None:
    try:
        result = float(value)
        return result if math.isfinite(result) else None
    except (TypeError, ValueError):
        return None


def _int(value: Any) -> int | None:
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _ra_to_deg(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        hours = float(value)
        return hours * 15.0 if abs(hours) <= 24 else hours
    text = str(value).strip()
    # Published catalogues often store sexagesimal coordinates compactly
    # (e.g. 123959 = 12h39m59s).  Do not treat those as decimal degrees.
    compact = re.fullmatch(r"(\d{2})(\d{2})(\d{2}(?:\.\d+)?)", text)
    if compact:
        hours = float(compact.group(1)) + float(compact.group(2)) / 60.0 + float(compact.group(3)) / 3600.0
        return hours * 15.0
    match = re.fullmatch(r"(\d{1,2})\s*:\s*(\d{1,2})(?:\s*:\s*([0-9.]+))?", text)
    if not match:
        return None
    hours = float(match.group(1)) + float(match.group(2)) / 60.0 + float(match.group(3) or 0) / 3600.0
    return hours * 15.0


def _parse_size(text: str | None) -> tuple[float | None, float | None]:
    """Best-effort parser; preserves original text even when ambiguous."""
    if not text:
        return None, None
    cleaned = text.lower().replace("×", "x").replace("deg", "°")
    numbers = re.findall(r"([0-9]+(?:\.[0-9]+)?)\s*(°|'|′|\")?", cleaned)
    if not numbers:
        return None, None
    values: list[float] = []
    for number, unit in numbers[:2]:
        value = float(number)
        if unit in ("'", "′"):
            value /= 60.0
        elif unit == '"':
            value /= 3600.0
        values.append(value)
    if len(values) == 1:
        return values[0], values[0]
    return values[0], values[1]


def _dec_to_deg(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("−", "-").replace("–", "-")
    # Compact DMS, e.g. -113723 = -11°37′23″.
    compact = re.fullmatch(r"([+-]?)(\d{2})(\d{2})(\d{2}(?:\.\d+)?)", text)
    if compact:
        sign = -1.0 if compact.group(1) == "-" else 1.0
        return sign * (float(compact.group(2)) + float(compact.group(3)) / 60.0 + float(compact.group(4)) / 3600.0)
    match = re.search(r"([+-]?)\s*(\d+(?:\.\d+)?)\s*[°d]?\s*(?:(\d+(?:\.\d+)?)\s*[\'′m])?", text)
    if not match:
        return None
    sign = -1.0 if match.group(1) == "-" else 1.0
    return sign * (float(match.group(2)) + float(match.group(3) or 0) / 60.0)


def _find_header_row(sheet, required: set[str], scan_rows: int = 30) -> tuple[int, dict[str, int]] | None:
    for row_number, row in enumerate(sheet.iter_rows(min_row=1, max_row=min(scan_rows, sheet.max_row), values_only=True), start=1):
        header_map = {str(v).strip(): i for i, v in enumerate(row) if v is not None}
        if required.issubset(header_map):
            return row_number, header_map
    return None


def preview_collection_import(path: str | Path) -> dict[str, Any]:
    path = Path(path)
    if path.suffix.casefold() == ".csv":
        table = discover_flexible_source(path)
        return {
            "format": "flexible", "label": "Unfamiliar catalogue — flexible import",
            "sheet": table["sheet"], "header_row": table["header_row"],
            "columns": table["headers"], "targets": count_flexible_rows(table, table["mapping"]),
            "mapping": table["mapping"],
        }
    from openpyxl import load_workbook
    workbook = load_workbook(path, read_only=True, data_only=True)
    for sheet in workbook.worksheets:
        james = _find_header_row(sheet, {"Catalog", "Common Name"})
        if james:
            row_no, headers = james
            count = sum(1 for row in sheet.iter_rows(min_row=row_no + 1, values_only=True) if any(v is not None for v in row))
            return {"format": "james", "label": "AstroFrame / James Imaging Guide", "sheet": sheet.title, "header_row": row_no, "columns": list(headers), "targets": count}
        bracken = _find_header_row(sheet, {"Object", "Rating", "Category", "Type", "Constellation", "RA", "Dec", "Size"})
        if bracken:
            row_no, headers = bracken
            count = sum(1 for row in sheet.iter_rows(min_row=row_no + 1, values_only=True) if row and row[headers["Object"]] is not None)
            return {"format": "bracken", "label": "Charles Bracken Highlight Objects", "sheet": sheet.title, "header_row": row_no, "columns": list(headers), "targets": count}
        gary = _find_header_row(sheet, {"Object Name & Image", "Type", "Size", "Rating", "Notes", "Right Ascension", "Declination", "Const."})
        if gary:
            row_no, headers = gary
            name_i = headers["Object Name & Image"]
            names = {_text(row[name_i]) for row in sheet.iter_rows(min_row=row_no + 3, values_only=True) if row and name_i < len(row) and _text(row[name_i])}
            count = len(names)
            return {
                "format": "gary_imm", "label": "Gary Imm Deep Sky Compendium",
                "sheet": sheet.title, "header_row": row_no,
                "columns": ["Object Name", "Type", "Subtype", "Class", "Size", "Distance", "Diameter", "Rating", "Notes", "RA", "Dec", "Constellation", "Nickname", "Alternate ID", "Nearby Objects", "Visual Magnitude", "Surface Brightness", "Inclination", "Priority", "Status"],
                "targets": count,
            }
    # RC22j: unfamiliar is no longer a dead end.  Return an inferred flexible
    # table so the UI can present a column-mapping screen.
    table = discover_flexible_source(path)
    return {
        "format": "flexible", "label": "Unfamiliar catalogue — flexible import",
        "sheet": table["sheet"], "header_row": table["header_row"],
        "columns": table["headers"], "targets": count_flexible_rows(table, table["mapping"]),
        "mapping": table["mapping"],
    }


def import_james_targets(path: str | Path, store: KnowledgeStore) -> CollectionRecord:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Spreadsheet import requires openpyxl. Install the updated requirements.txt first.") from exc

    path = Path(path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = next((ws for ws in workbook.worksheets if _find_header_row(ws, {"Catalog", "Common Name"})), None)
    if sheet is None:
        raise ValueError("This workbook does not contain a recognised AstroFrame target table.")
    header_row, header_map = _find_header_row(sheet, {"Catalog", "Common Name"})
    rows = sheet.iter_rows(min_row=header_row + 1, values_only=True)

    def value(row: tuple[Any, ...], column: str) -> Any:
        index = header_map.get(column)
        return row[index] if index is not None and index < len(row) else None

    entries: list[CollectionEntry] = []
    seen: set[str] = set()
    for row in rows:
        catalog = _text(value(row, "Catalog"))
        common = _text(value(row, "Common Name"))
        if not catalog and not common:
            continue
        canonical = catalog or common or "Unknown target"
        aliases = [a for a in [common] if a and a.casefold() != canonical.casefold()]
        size_text = _text(value(row, "Apparent Size"))
        width_deg, height_deg = _parse_size(size_text)
        target = TargetRecord(
            id=target_id_for(canonical),
            canonical_name=canonical,
            common_name=common,
            aliases=aliases,
            ra_deg=_ra_to_deg(value(row, "RA")),
            dec_deg=_float(value(row, "Dec (deg)")),
            angular_width_deg=width_deg,
            angular_height_deg=height_deg,
            apparent_size_text=size_text,
            object_type=_text(value(row, "Type")),
            constellation=_text(value(row, "Constellation")),
        )
        target = store.upsert_target(target)
        if target.id in seen:
            continue
        seen.add(target.id)

        source_fields = {
            "magnitude": _text(value(row, "Mag")),
            "suggested_focal_length_full_frame": _text(value(row, "Suggested FL (full-frame)")),
            "max_altitude_source": _float(value(row, "Max Alt @ your lat")),
        }
        source_fields = {k: v for k, v in source_fields.items() if v is not None}
        entries.append(CollectionEntry(
            target_id=target.id,
            source_name=common or catalog,
            rank=_int(value(row, "Rank")),
            tier=_text(value(row, "Tier")),
            difficulty=_text(value(row, "Level")),
            fov_class=_text(value(row, "Best Field")),
            narrowband=_text(value(row, "Narrowband")),
            broadband=_text(value(row, "Broadband")),
            sho=_text(value(row, "SHO")),
            hoo=_text(value(row, "HOO")),
            moon_ok=_text(value(row, "Moon OK")),
            best_month=_text(value(row, "Best Month (evening)")),
            visibility=_text(value(row, "Visibility")),
            notes=_text(value(row, "Notes")),
            source_fields=source_fields,
        ))

    collection = CollectionRecord(
        id="james_imaging_guide",
        name="James' Imaging Guide",
        author="James Glucksman",
        description="Curated astrophotography targets imported from astrophotography_targets.xlsx.",
        version="1",
        source_type="user",
        imported_from=path.name,
        entries=entries,
    )
    store.save_collection(collection)
    return collection


def import_bracken_targets(path: str | Path, store: KnowledgeStore) -> CollectionRecord:
    from openpyxl import load_workbook
    path = Path(path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = next((ws for ws in workbook.worksheets if _find_header_row(ws, {"Object", "Rating", "Category", "Type", "Constellation", "RA", "Dec", "Size"})), None)
    if sheet is None:
        raise ValueError("This workbook does not contain a recognised Charles Bracken target table.")
    header_row, header_map = _find_header_row(sheet, {"Object", "Rating", "Category", "Type", "Constellation", "RA", "Dec", "Size"})
    def value(row, column):
        i = header_map.get(column); return row[i] if i is not None and i < len(row) else None
    entries = []
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        raw_name = _text(value(row, "Object"))
        if not raw_name: continue
        parts = [p.strip() for p in raw_name.split(",", 1)]
        canonical = parts[0]; common = parts[1] if len(parts) > 1 else None
        aliases = [common] if common else []
        size_text = _text(value(row, "Size")); width, height = _parse_size(size_text)
        target = store.upsert_target(TargetRecord(id=target_id_for(canonical), canonical_name=canonical, common_name=common, aliases=aliases, ra_deg=_ra_to_deg(value(row, "RA")), dec_deg=_dec_to_deg(value(row, "Dec")), angular_width_deg=width, angular_height_deg=height, apparent_size_text=size_text, object_type=_text(value(row, "Type")), constellation=_text(value(row, "Constellation"))))
        source_fields = {"category": _text(value(row, "Category")), "opposition": _text(value(row, "Opposition")), "hemisphere": _text(value(row, "Hemisphere")), "size": size_text}
        source_fields = {k:v for k,v in source_fields.items() if v is not None}
        entries.append(CollectionEntry(target_id=target.id, source_name=raw_name, rating=_float(value(row, "Rating")), source_fields=source_fields))
    collection = CollectionRecord(id="charles_bracken_highlight_objects", name="Charles Bracken's Highlight Objects", author="Charles Bracken", description="Highlight objects from The Astrophotographer's Universe, preserving source metadata including the added Hemisphere field when present.", version="1", source_type="published", imported_from=path.name, entries=entries)
    store.save_collection(collection); return collection


def import_gary_imm_targets(path: str | Path, store: KnowledgeStore) -> CollectionRecord:
    """Import the multi-row Main table from Gary Imm's Deep Sky Compendium.

    The workbook is left untouched.  AstroFrame maps the source's three-row
    heading structure internally and retains useful source-only metadata.
    """
    from openpyxl import load_workbook
    path = Path(path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    found = None
    for sheet in workbook.worksheets:
        match = _find_header_row(sheet, {"Object Name & Image", "Type", "Size", "Rating", "Notes", "Right Ascension", "Declination", "Const."})
        if match:
            found = (sheet, match[0], match[1]); break
    if found is None:
        raise ValueError("This workbook does not contain a recognised Gary Imm Deep Sky Compendium table.")
    sheet, header_row, h = found

    # Gary's Main sheet uses three heading rows.  Several important columns
    # have blank top-level headings, so their positions are intentionally
    # anchored relative to the named RA/Dec headings rather than guessed from
    # the user's data.
    def idx(name: str) -> int: return h[name]
    name_i=idx("Object Name & Image"); type_i=idx("Type"); sub_i=idx("Sub"); class_i=idx("Class")
    size_i=idx("Size"); distance_i=idx("Distance"); diameter_i=idx("Diameter"); rating_i=idx("Rating"); notes_i=idx("Notes")
    ra_hms_i=idx("Right Ascension"); ra_deg_i=ra_hms_i+1
    dec_dms_i=idx("Declination"); dec_deg_i=dec_dms_i+1
    const_i=idx("Const."); nick_i=idx("Nick."); alt_i=idx("Alt. ID"); nearby_i=idx("Nearby")
    visual_i=idx("Visual"); surf_i=idx("Surf."); incl_i=idx("Inclin.")
    my_cols=[i for i,v in enumerate(sheet[header_row]) if _text(v.value)=="My"]
    priority_i=my_cols[0] if len(my_cols)>0 else None; status_i=my_cols[1] if len(my_cols)>1 else None; mynotes_i=my_cols[2] if len(my_cols)>2 else None

    def val(row, i): return row[i] if i is not None and i < len(row) else None
    def split_aliases(text):
        t=_text(text)
        if not t: return []
        return [x.strip() for x in re.split(r"[,;/]", t) if x.strip() and x.strip() not in {"-", "—"}]

    entries=[]; seen=set()
    # Data begins after the three heading rows (7, 8 and 9 in the 2026 book).
    for row in sheet.iter_rows(min_row=header_row+3, values_only=True):
        canonical=_text(val(row,name_i))
        if not canonical: continue
        nickname=_text(val(row,nick_i)); alt=_text(val(row,alt_i))
        aliases=[]
        if nickname and nickname not in {"-", "—"}: aliases.append(nickname)
        aliases.extend(split_aliases(alt))
        size_raw=_float(val(row,size_i)); size_text=_text(val(row,size_i))
        size_deg=(size_raw/60.0) if size_raw is not None else None
        # Prefer the full-precision sexagesimal source columns.  The adjacent
        # decimal columns in Gary Imm's workbook are display-rounded (M104 is
        # 190.0, -11.6 there), which is far too coarse for image markers.
        ra_deg=_ra_to_deg(val(row,ra_hms_i))
        dec_deg=_dec_to_deg(val(row,dec_dms_i))
        if ra_deg is None: ra_deg=_float(val(row,ra_deg_i))
        if dec_deg is None: dec_deg=_float(val(row,dec_deg_i))
        incoming=TargetRecord(
            id=target_id_for(canonical), canonical_name=canonical,
            common_name=nickname if nickname not in {"-", "—"} else None,
            aliases=aliases, ra_deg=ra_deg, dec_deg=dec_deg,
            angular_width_deg=size_deg, angular_height_deg=size_deg,
            apparent_size_text=(f"{size_raw:g}′" if size_raw is not None else size_text),
            object_type=_text(val(row,type_i)), constellation=_text(val(row,const_i)),
        )
        # Gary contains thousands of rows.  New IDs can be inserted directly;
        # only genuine ID collisions need the more expensive merge path.
        if incoming.id in store._targets:
            target=store.upsert_target(incoming, save=False)
        else:
            store._targets[incoming.id]=incoming
            target=incoming
        if target.id in seen: continue
        seen.add(target.id)
        source_fields={
            "subtype":_text(val(row,sub_i)), "class":_text(val(row,class_i)),
            "distance":_text(val(row,distance_i)), "diameter":_text(val(row,diameter_i)),
            "alternate_id":alt, "nearby_objects":_text(val(row,nearby_i)),
            "visual_magnitude":_text(val(row,visual_i)), "surface_brightness":_text(val(row,surf_i)),
            "inclination_deg":_float(val(row,incl_i)), "priority":_text(val(row,priority_i)),
            "status":_text(val(row,status_i)), "my_notes":_text(val(row,mynotes_i)),
            "ra_source":_text(val(row,ra_hms_i)), "dec_source":_text(val(row,dec_dms_i)),
        }
        source_fields={k:v for k,v in source_fields.items() if v is not None and v != " "}
        entries.append(CollectionEntry(
            target_id=target.id, source_name=canonical,
            rating=_float(val(row,rating_i)), notes=_text(val(row,notes_i)),
            source_fields=source_fields,
        ))
    # One disk write after the large import rather than thousands of writes.
    store._save_targets()
    collection=CollectionRecord(
        id="gary_imm_deep_sky_compendium", name="Gary Imm's Deep Sky Compendium",
        author="Gary Imm", description="Deep-sky compendium imported directly from Gary Imm's original multi-row workbook; useful source metadata is preserved without requiring reformatting.",
        version="2026", source_type="published", imported_from=path.name, entries=entries,
    )
    store.save_collection(collection)
    return collection


# ---------------- Flexible catalogue import (RC22j) ----------------

FLEXIBLE_FIELDS = [
    ("name", "Primary name", True),
    ("common_name", "Common name", False),
    ("aliases", "Other names / aliases", False),
    ("object_type", "Object type", False),
    ("constellation", "Constellation", False),
    ("magnitude", "Magnitude", False),
    ("surface_brightness", "Surface brightness", False),
    ("width", "Major / width", False),
    ("height", "Minor / height", False),
    ("position_angle", "Position angle", False),
    ("best_month", "Best month", False),
    ("ra", "RA (single column)", False),
    ("ra_h", "RA hours", False),
    ("ra_m", "RA minutes", False),
    ("ra_s", "RA seconds", False),
    ("dec", "Dec (single column)", False),
    ("dec_sign", "Dec sign", False),
    ("dec_d", "Dec degrees", False),
    ("dec_m", "Dec minutes", False),
    ("dec_s", "Dec seconds", False),
]


def _normalise_header(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"[^a-z0-9]+", " ", str(value).strip().casefold()).strip()


def _dedupe_headers(values: list[Any], fallback_values: list[Any] | None = None) -> list[str]:
    headers: list[str] = []
    seen: dict[str, int] = {}
    fallback_values = fallback_values or []
    for i, value in enumerate(values):
        raw = _text(value)
        if not raw and i < len(fallback_values):
            raw = _text(fallback_values[i])
        raw = (raw or f"Column {i + 1}").replace("\n", " ").strip()
        count = seen.get(raw.casefold(), 0) + 1
        seen[raw.casefold()] = count
        headers.append(raw if count == 1 else f"{raw} ({count})")
    return headers


def _row_has_data(row: tuple[Any, ...] | list[Any]) -> bool:
    return any(_text(v) is not None for v in row)


def _read_flexible_table(path: str | Path, *, sheet_name: str | None = None, header_row: int | None = None) -> dict[str, Any]:
    """Read an unfamiliar xlsx/xlsm/csv into a simple header/row structure.

    This deliberately avoids modifying the source workbook.  The second row is
    used only to fill blank top-level headings, which handles real-world sheets
    such as BAM600 without forcing a special-case importer.
    """
    path = Path(path)
    suffix = path.suffix.casefold()
    if suffix == ".csv":
        import csv
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            raw_rows = list(csv.reader(handle))
        if not raw_rows:
            raise ValueError("This CSV is empty.")
        hdr = max(1, int(header_row or 1))
        if hdr > len(raw_rows):
            raise ValueError("Header row is beyond the end of the CSV.")
        top = raw_rows[hdr - 1]
        headers = _dedupe_headers(top)
        rows = [tuple(r) for r in raw_rows[hdr:] if _row_has_data(r)]
        return {"sheet": "CSV", "sheet_names": ["CSV"], "header_row": hdr, "headers": headers, "rows": rows}

    if suffix not in {".xlsx", ".xlsm"}:
        raise ValueError("Flexible import currently supports .xlsx, .xlsm and .csv files.")
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    sheet_names = list(wb.sheetnames)
    if not sheet_names:
        raise ValueError("This workbook contains no worksheets.")
    ws = wb[sheet_name] if sheet_name in sheet_names else wb[sheet_names[0]]
    hdr = max(1, int(header_row or 1))
    if hdr > ws.max_row:
        raise ValueError("Header row is beyond the end of the selected worksheet.")
    top = [cell.value for cell in next(ws.iter_rows(min_row=hdr, max_row=hdr))]
    second = []
    if hdr + 1 <= ws.max_row:
        second = [cell.value for cell in next(ws.iter_rows(min_row=hdr + 1, max_row=hdr + 1))]
    headers = _dedupe_headers(top, second)
    rows = [tuple(r) for r in ws.iter_rows(min_row=hdr + 1, values_only=True) if _row_has_data(r)]
    return {"sheet": ws.title, "sheet_names": sheet_names, "header_row": hdr, "headers": headers, "rows": rows}


def _score_header(header: str, patterns: list[str]) -> int:
    norm = _normalise_header(header)
    if not norm:
        return 0
    score = 0
    for pattern in patterns:
        if norm == pattern:
            score = max(score, 100)
        elif pattern in norm:
            score = max(score, 60)
        elif all(part in norm.split() for part in pattern.split()):
            score = max(score, 45)
    return score


def infer_flexible_mapping(headers: list[str]) -> dict[str, int | None]:
    patterns = {
        "name": ["target name", "object name", "object", "name", "catalog", "catalogue"],
        "common_name": ["common name", "nickname", "nick"],
        "aliases": ["other names", "other name", "alternate id", "alt id", "aliases", "alias"],
        "object_type": ["object type", "type"],
        "constellation": ["constellation", "const", "con"],
        "magnitude": ["vmag", "visual magnitude", "magnitude", "mag"],
        "surface_brightness": ["surface brightness", "surf", "sb"],
        "width": ["major axis", "major", "max", "width", "size"],
        "height": ["minor axis", "minor", "min", "height"],
        "position_angle": ["position angle", "pa"],
        "best_month": ["best month", "month to observe"],
        "ra": ["right ascension", "ra", "raj2000", "ra j2000"],
        "ra_h": ["ra2k", "ra hours", "ra hour", "rah"],
        "ra_m": ["ra min", "ra minutes", "ram"],
        "ra_s": ["ra sec", "ra seconds", "ras"],
        "dec": ["declination", "dec", "dej2000", "dec j2000"],
        "dec_sign": ["dec sign", "declination sign"],
        "dec_d": ["dec deg", "dec degrees", "declination degrees", "ded"],
        "dec_m": ["dec min", "dec minutes", "declination minutes", "dem"],
        "dec_s": ["dec sec", "dec seconds", "declination seconds", "des"],
    }
    mapping: dict[str, int | None] = {key: None for key, _label, _required in FLEXIBLE_FIELDS}
    used: set[int] = set()
    for key, pats in patterns.items():
        ranked = sorted((( _score_header(h, pats), i) for i, h in enumerate(headers)), reverse=True)
        if ranked and ranked[0][0] >= 45 and ranked[0][1] not in used:
            mapping[key] = ranked[0][1]
            used.add(ranked[0][1])

    # Common multi-component shorthand used by observing lists.  Header names
    # like RA2k/min/sec and dec/d/m/s are otherwise too terse to infer safely.
    norms = [_normalise_header(h) for h in headers]
    # _dedupe_headers appends (2), (3), ... to repeated headings for display.
    # Strip that numeric suffix for inference so RA "min" is not mistaken
    # for the later Dec "m" simply because a size column was also called Min.
    inference_norms = [re.sub(r"\s+\d+$", "", n) for n in norms]
    def first_exact(options: set[str], *, after: int = -1) -> int | None:
        for i, n in enumerate(inference_norms):
            if i > after and n in options:
                return i
        return None
    if mapping["ra_h"] is None:
        mapping["ra_h"] = first_exact({"ra2k", "ra h", "rah"})
    if mapping["ra_h"] is not None:
        h = int(mapping["ra_h"])
        mapping["ra_m"] = mapping["ra_m"] if mapping["ra_m"] is not None else first_exact({"min", "m", "ra min"}, after=h)
        after_m = int(mapping["ra_m"]) if mapping["ra_m"] is not None else h
        mapping["ra_s"] = mapping["ra_s"] if mapping["ra_s"] is not None else first_exact({"sec", "s", "ra sec"}, after=after_m)
    if mapping["dec_sign"] is None:
        mapping["dec_sign"] = first_exact({"dec", "sign", "dec sign"})
    if mapping["dec_sign"] is not None:
        d0 = int(mapping["dec_sign"])
        mapping["dec_d"] = mapping["dec_d"] if mapping["dec_d"] is not None else first_exact({"d", "deg", "degrees"}, after=d0)
        d1 = int(mapping["dec_d"]) if mapping["dec_d"] is not None else d0
        mapping["dec_m"] = mapping["dec_m"] if mapping["dec_m"] is not None else first_exact({"m", "min", "minutes"}, after=d1)
        d2 = int(mapping["dec_m"]) if mapping["dec_m"] is not None else d1
        mapping["dec_s"] = mapping["dec_s"] if mapping["dec_s"] is not None else first_exact({"s", "sec", "seconds"}, after=d2)

    # Prefer component coordinates when all parts were detected.
    if all(mapping[k] is not None for k in ("ra_h", "ra_m", "ra_s")):
        mapping["ra"] = None
    if mapping["dec_d"] is not None and mapping["dec_m"] is not None:
        mapping["dec"] = None
    return mapping


def discover_flexible_source(path: str | Path) -> dict[str, Any]:
    """Choose the most table-like worksheet/header row and infer a mapping."""
    path = Path(path)
    if path.suffix.casefold() == ".csv":
        table = _read_flexible_table(path, header_row=1)
        table["mapping"] = infer_flexible_mapping(table["headers"])
        return table
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    best: tuple[float, dict[str, Any]] | None = None
    vocabulary = {"name", "object", "target", "ra", "dec", "type", "constellation", "const", "con", "vmag", "mag", "max", "min", "pa"}
    for ws in wb.worksheets:
        max_scan = min(ws.max_row, 20)
        for row_no, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1):
            values = [_text(v) for v in row]
            nonempty = [v for v in values if v]
            if len(nonempty) < 2:
                continue
            words = set()
            for value in nonempty[:80]:
                words.update(_normalise_header(value).split())
            lexical = len(words & vocabulary)
            score = lexical * 12 + min(len(nonempty), 30)
            try:
                table = _read_flexible_table(path, sheet_name=ws.title, header_row=row_no)
                mapping = infer_flexible_mapping(table["headers"])
                mapped = sum(v is not None for v in mapping.values())
                score += mapped * 8
                # A plausible table needs a name and some form of coordinates.
                if mapping.get("name") is not None:
                    score += 45
                if mapping.get("ra") is not None or mapping.get("ra_h") is not None:
                    score += 25
                if mapping.get("dec") is not None or mapping.get("dec_d") is not None:
                    score += 25
                if best is None or score > best[0]:
                    table["mapping"] = mapping
                    best = (score, table)
            except Exception:
                continue
    if best is None:
        raise ValueError("AstroFrame could not find a plausible target table in this workbook.")
    return best[1]


def _cell(row: tuple[Any, ...], index: int | None) -> Any:
    return row[index] if index is not None and 0 <= index < len(row) else None


def _component_ra(row: tuple[Any, ...], mapping: dict[str, int | None]) -> float | None:
    single = _ra_to_deg(_cell(row, mapping.get("ra")))
    if single is not None:
        return single
    h = _float(_cell(row, mapping.get("ra_h")))
    m = _float(_cell(row, mapping.get("ra_m"))) or 0.0
    s = _float(_cell(row, mapping.get("ra_s"))) or 0.0
    if h is None:
        return None
    return (h + m / 60.0 + s / 3600.0) * 15.0


def _component_dec(row: tuple[Any, ...], mapping: dict[str, int | None]) -> float | None:
    single = _dec_to_deg(_cell(row, mapping.get("dec")))
    if single is not None:
        return single
    d = _float(_cell(row, mapping.get("dec_d")))
    if d is None:
        return None
    m = _float(_cell(row, mapping.get("dec_m"))) or 0.0
    s = _float(_cell(row, mapping.get("dec_s"))) or 0.0
    sign_raw = _text(_cell(row, mapping.get("dec_sign"))) or ""
    sign = -1.0 if sign_raw.strip().startswith(("-", "−", "–")) or d < 0 else 1.0
    return sign * (abs(d) + m / 60.0 + s / 3600.0)


def _dimension_deg(value: Any, unit: str = "arcmin") -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        v = _float(value)
        if v is None:
            return None
        if unit == "degrees":
            return v
        if unit == "arcsec":
            return v / 3600.0
        return v / 60.0
    text = str(value).strip()
    if not text:
        return None
    # Explicit unit markers override the mapper's default.
    if '"' in text or "arcsec" in text.casefold():
        number = _float(re.sub(r"[^0-9.+-]", "", text))
        return number / 3600.0 if number is not None else None
    if "'" in text or "′" in text or "arcmin" in text.casefold():
        number = _float(re.sub(r"[^0-9.+-]", "", text))
        return number / 60.0 if number is not None else None
    if "°" in text or "deg" in text.casefold():
        number = _float(re.sub(r"[^0-9.+-]", "", text))
        return number
    return _dimension_deg(_float(text), unit)


def _split_aliases_flexible(value: Any) -> list[str]:
    text = _text(value)
    if not text:
        return []
    return [part.strip() for part in re.split(r"[,;/|]", text) if part.strip() and part.strip() not in {"-", "—"}]


def flexible_preview_rows(table: dict[str, Any], mapping: dict[str, int | None], *, size_unit: str = "arcmin", limit: int = 5) -> list[dict[str, Any]]:
    preview: list[dict[str, Any]] = []
    for row in table["rows"]:
        name = _text(_cell(row, mapping.get("name")))
        if not name:
            continue
        ra = _component_ra(row, mapping)
        dec = _component_dec(row, mapping)
        if ra is None or dec is None:
            continue
        preview.append({
            "name": name,
            "common_name": _text(_cell(row, mapping.get("common_name"))),
            "type": _text(_cell(row, mapping.get("object_type"))),
            "constellation": _text(_cell(row, mapping.get("constellation"))),
            "ra_deg": ra,
            "dec_deg": dec,
            "width_deg": _dimension_deg(_cell(row, mapping.get("width")), size_unit),
            "height_deg": _dimension_deg(_cell(row, mapping.get("height")), size_unit),
        })
        if len(preview) >= limit:
            break
    return preview


def count_flexible_rows(table: dict[str, Any], mapping: dict[str, int | None]) -> int:
    count = 0
    for row in table["rows"]:
        if _text(_cell(row, mapping.get("name"))) and _component_ra(row, mapping) is not None and _component_dec(row, mapping) is not None:
            count += 1
    return count


def import_flexible_collection(
    path: str | Path, store: KnowledgeStore, *, table: dict[str, Any], mapping: dict[str, int | None],
    collection_name: str, author: str | None = None, size_unit: str = "arcmin"
) -> CollectionRecord:
    if mapping.get("name") is None:
        raise ValueError("Map a Primary name column before importing.")
    entries: list[CollectionEntry] = []
    headers = table["headers"]
    for row in table["rows"]:
        canonical = _text(_cell(row, mapping.get("name")))
        if not canonical:
            continue
        ra_deg = _component_ra(row, mapping)
        dec_deg = _component_dec(row, mapping)
        if ra_deg is None or dec_deg is None:
            continue
        common = _text(_cell(row, mapping.get("common_name")))
        aliases = _split_aliases_flexible(_cell(row, mapping.get("aliases")))
        if common and common.casefold() != canonical.casefold() and common not in aliases:
            aliases.append(common)
        width = _dimension_deg(_cell(row, mapping.get("width")), size_unit)
        height = _dimension_deg(_cell(row, mapping.get("height")), size_unit)
        if width is not None and height is None:
            height = width
        if height is not None and width is None:
            width = height
        size_text = None
        raw_w = _text(_cell(row, mapping.get("width")))
        raw_h = _text(_cell(row, mapping.get("height")))
        if raw_w or raw_h:
            size_text = " × ".join(v for v in (raw_w, raw_h) if v)
        incoming = TargetRecord(
            id=target_id_for(canonical), canonical_name=canonical, common_name=common, aliases=aliases,
            ra_deg=ra_deg, dec_deg=dec_deg, angular_width_deg=width, angular_height_deg=height,
            apparent_size_text=size_text, position_angle_deg=_float(_cell(row, mapping.get("position_angle"))),
            object_type=_text(_cell(row, mapping.get("object_type"))), constellation=_text(_cell(row, mapping.get("constellation"))),
        )
        target = store.upsert_target(incoming, save=False) if incoming.id in store._targets else incoming
        if incoming.id not in store._targets:
            store._targets[incoming.id] = incoming
        # Preserve every source-list row, even when two rows intentionally
        # describe the same physical object under different catalogue names.
        # Knowledge targets may merge aliases, but a BAM600-style observing
        # list must still remain a 600-entry list.
        source_fields: dict[str, Any] = {}
        for i, header in enumerate(headers):
            value = _cell(row, i)
            if _text(value) is not None:
                source_fields[header] = value
        entry = CollectionEntry(
            target_id=target.id, source_name=canonical,
            best_month=_text(_cell(row, mapping.get("best_month"))),
            source_fields=source_fields,
        )
        mag = _float(_cell(row, mapping.get("magnitude")))
        sb = _float(_cell(row, mapping.get("surface_brightness")))
        if mag is not None:
            entry.source_fields["magnitude"] = mag
        if sb is not None:
            entry.source_fields["surface_brightness"] = sb
        entries.append(entry)
    if not entries:
        raise ValueError("No rows had a usable name, RA and Dec with the current mapping.")
    store._save_targets()
    collection = CollectionRecord(
        id=slugify(collection_name), name=collection_name.strip() or Path(path).stem, author=_text(author),
        description=f"Imported with AstroFrame's flexible catalogue mapper from {Path(path).name} ({table['sheet']}).",
        version="1", source_type="user", imported_from=Path(path).name, entries=entries,
    )
    store.save_collection(collection)
    return collection

def import_target_collection(path: str | Path, store: KnowledgeStore) -> CollectionRecord:
    preview = preview_collection_import(path)
    if preview["format"] == "james": return import_james_targets(path, store)
    if preview["format"] == "bracken": return import_bracken_targets(path, store)
    if preview["format"] == "gary_imm": return import_gary_imm_targets(path, store)
    raise ValueError("Recognised spreadsheet format has no importer.")
